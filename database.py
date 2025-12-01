
import pandas as pd
import os
from openpyxl import load_workbook, Workbook

DB_FILE = "coffeeshop_database.xlsx"
SHEET_MENU = "Menu_Costing"
SHEET_INVENTORY = "Inventory_Stock"
CURRENCY = "₱"

def load_sheet(sheet_name):
    """Loads a specific sheet from the Excel file into a DataFrame."""
    if not os.path.exists(DB_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_excel(DB_FILE, sheet_name=sheet_name)
        return df
    except ValueError:
        # Sheet does not exist, return empty DataFrame
        return pd.DataFrame()
    except Exception as e:
        print(f"Error loading {sheet_name}: {e}")
        return pd.DataFrame()

def save_dataframe(df, sheet_name, mode='append'):
    """Saves/appends a DataFrame to a specific sheet in the Excel file and auto-adjusts columns."""

    if mode == 'append':
        existing_df = load_sheet(sheet_name)
        if not existing_df.empty:
            df = pd.concat([existing_df, df], ignore_index=True)

    try:
        # Try to load the file
        book = load_workbook(DB_FILE)
    except Exception:
        # File is corrupted or doesn't exist, create new
        book = Workbook()

    # Remove all sheets except the current one and SHEET_MENU
    for sheet in list(book.sheetnames):
        if sheet != sheet_name and sheet != SHEET_MENU:
            book.remove(book[sheet])

    # Remove the sheet if it exists (in case it's the current one)
    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])

    # Create a new sheet with the data
    sheet = book.create_sheet(sheet_name)
    # Write the DataFrame to the sheet
    for r, (idx, row) in enumerate(df.iterrows(), start=1):
        for c, (col_name, value) in enumerate(row.items(), start=1):
            if r == 1:  # Write headers
                sheet.cell(row=r, column=c, value=col_name)
            sheet.cell(row=r+1, column=c, value=value)

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3)
        sheet.column_dimensions[column].width = adjusted_width

    # Remove default sheet if it's empty and not needed
    if 'Sheet' in book.sheetnames and len(book['Sheet']) == 0:
        book.remove(book['Sheet'])

    book.save(DB_FILE)
    return df

def update_inventory(ingredient_name, quantity_deducted):
    """Deducts a quantity from the inventory stock for a specific ingredient."""
    df_inventory = load_sheet(SHEET_INVENTORY)
    
    if df_inventory.empty:
        print("Inventory is empty. Cannot deduct stock.")
        return False
        
    # Find the ingredient row
    row_index = df_inventory[df_inventory['Ingredient Name'] == ingredient_name].index
    
    if not row_index.empty:
        current_stock = df_inventory.loc[row_index[0], 'Current Stock (g/ml)']
        new_stock = current_stock - quantity_deducted
        
        # Ensure stock doesn't go negative (though it could for a backorder situation)
        df_inventory.loc[row_index[0], 'Current Stock (g/ml)'] = max(0, new_stock)
        
        # Save the updated DataFrame back to the Excel file
        save_dataframe(df_inventory, SHEET_INVENTORY, mode='overwrite')
        return True
    else:
        print(f"Ingredient '{ingredient_name}' not found in inventory.")
        return False